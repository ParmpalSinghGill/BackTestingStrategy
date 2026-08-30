"""
ProcessPool Multi-Experiment Runner for Swing Strategy

Executes 6 Capital x Risk Cap Scenarios Simultaneously across 6 CPU Processes (ProcessPoolExecutor):
1. 50k Capital | 1.0k Risk Cap
2. 50k Capital | 0.5k Risk Cap
3. 100k Capital | 1.0k Risk Cap
4. 100k Capital | 0.5k Risk Cap
5. 200k Capital | 1.0k Risk Cap
6. 200k Capital | 0.5k Risk Cap

For each scenario, generates isolated report and plot folders containing:
- Excel Account Statement with Active Position Count & native clickable =HYPERLINK(...) trade plot links
- Trade Candlestick PNG plot charts in trade_charts/
- Monthly Returns Heatmap PNG
- Yearly Returns Breakdown PNG
- Monthly Capital Growth PNG
- Interactive Live Hover Equity Curve HTML

Exports Master_Experiments_Comparison.xlsx synthesizing all results side-by-side.
"""

import os
import sys
from pathlib import Path
import pandas as pd
from concurrent.futures import ProcessPoolExecutor

# Force UTF-8 encoding for stdout on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

MULTI_REPORTS_DIR = BASE_DIR / "Reports" / "Multi_Experiment_Suite"
MULTI_PLOTS_DIR = BASE_DIR / "Plots" / "Multi_Experiment_Suite"

os.makedirs(MULTI_REPORTS_DIR, exist_ok=True)
os.makedirs(MULTI_PLOTS_DIR, exist_ok=True)

from swing_strategy.strategy_engine import prepare_swing_strategy_dataset, run_swing_strategy_ml_model
from swing_strategy.generate_statement import generate_swing_strategy_statement


def run_single_scenario_proc(sc):
    exp_name = sc["exp_name"]
    deposit = sc["initial_deposit"]
    risk = sc["max_risk"]

    rep_dir = MULTI_REPORTS_DIR / exp_name
    plt_dir = MULTI_PLOTS_DIR / exp_name

    res_summary = generate_swing_strategy_statement(
        initial_deposit=deposit,
        max_charts_to_generate=20,
        max_risk_per_trade=risk,
        df_acc=None,
        custom_reports_dir=rep_dir,
        custom_plots_dir=plt_dir,
        exp_name=exp_name
    )
    return res_summary


def run_all_experiments():
    print("==========================================================================", flush=True)
    print("      SWING STRATEGY BATCH MULTI-EXPERIMENT MATRIX SUITE                  ", flush=True)
    print("==========================================================================", flush=True)

    # Scenarios matrix definition
    scenarios = [
        {"exp_name": "Exp_50k_1.0k", "initial_deposit": 50000.0, "max_risk": 1000.0},
        {"exp_name": "Exp_50k_0.5k", "initial_deposit": 50000.0, "max_risk": 500.0},
        {"exp_name": "Exp_100k_1.0k", "initial_deposit": 100000.0, "max_risk": 1000.0},
        {"exp_name": "Exp_100k_0.5k", "initial_deposit": 100000.0, "max_risk": 500.0},
        {"exp_name": "Exp_200k_1.0k", "initial_deposit": 200000.0, "max_risk": 1000.0},
        {"exp_name": "Exp_200k_0.5k", "initial_deposit": 200000.0, "max_risk": 500.0},
    ]

    print("\n--- Launching 6 CPU Core Processes SIMULTANEOUSLY in Parallel ---\n", flush=True)

    with ProcessPoolExecutor(max_workers=6) as executor:
        futures = [executor.submit(run_single_scenario_proc, sc) for sc in scenarios]
        summary_list = [f.result() for f in futures]

    summary_list.sort(key=lambda x: [sc["exp_name"] for sc in scenarios].index(x["Exp_Name"]))

    df_master = pd.DataFrame(summary_list)

    master_excel_path = MULTI_REPORTS_DIR / "Master_Experiments_Comparison.xlsx"
    master_csv_path = MULTI_REPORTS_DIR / "Master_Experiments_Comparison.csv"

    # Formatting summary display table
    display_rows = []
    for s in summary_list:
        display_rows.append({
            "Experiment ID": s["Exp_Name"],
            "Initial Capital": f"Rs {s['Initial_Capital']:,.0f}",
            "Risk Cap / Trade": f"Rs {s['Max_Risk_Cap']:,.0f}",
            "Final Account Equity": f"Rs {s['Final_Equity']:,.2f}",
            "Net Return (%)": f"+{s['Total_Net_Return_Pct']:,.2f}%",
            "CAGR (%)": f"{s['CAGR_Pct']:.2f}%",
            "Executed Trades": f"{s['Executed_Trades']:,}",
            "Win Rate (%)": f"{s['Win_Rate_Pct']:.2f}%",
            "Max Drawdown (%)": f"{s['Max_Drawdown_Pct']:.2f}%",
            "Taxes Paid": f"Rs {s['Total_Taxes_Paid']:,.2f}"
        })
    df_display = pd.DataFrame(display_rows)

    print("\n==========================================================================================", flush=True)
    print("                         MASTER EXPERIMENTS COMPARISON SUMMARY                            ", flush=True)
    print("==========================================================================================", flush=True)
    print(df_display.to_string(index=False), flush=True)
    print("==========================================================================================\n", flush=True)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scenarios Comparison"

        headers = list(df_display.columns)
        ws.append(headers)

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, row in df_display.iterrows():
            row_data = list(row)
            ws.append(row_data)

        wb.save(master_excel_path)
        print(f"Master Comparison Excel Report saved to: {master_excel_path.resolve()}", flush=True)
    except Exception as e:
        print(f"Master Excel export warning: {e}", flush=True)

    df_display.to_csv(master_csv_path, index=False)
    return master_excel_path, df_display


if __name__ == "__main__":
    run_all_experiments()
