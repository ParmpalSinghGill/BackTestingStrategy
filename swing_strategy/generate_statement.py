"""
Swing Strategy Account Statement Generator (True Realistic Fills Model)

Initial Capital Deposit: Configurable via initial_deposit parameter
Max Loss Risk Cap: Configurable via max_risk_per_trade parameter
Execution Rules & Order of Operations:
1. ENTRIES FIRST: On each trading day D, new entries process FIRST using cash available at the start of Day D.
2. EXITS SECOND: Open trades whose exit date <= D process SECOND. Cash freed by Day D exits is available on Day D+1 onwards.
3. Gap-Up Entry: If C3 Open > Entry Price, buys at C3 Open * 1.002 (0.2% above C3 Open).
4. Gap Exit: If Exit Open > Target Price or Exit Open < SL Price, sells at Exit Open * 0.999 (0.1% below Exit Open).
5. Sizing: Quantity = max(1, int(1000.0 / (Entry_Price - SL_Price))).
6. Active Position Count: Tracks exact count of open active trades on every transaction.

Generates:
1. Chronological Buy/Exit trade log with running capital balance updates and active position counts.
2. High-resolution PNG candlestick charts saved to Plots/swing_statement_trades/.
3. Formatted Excel Account Statement exported to Reports/Swing_Strategy_Account_Statement_{cap}k.xlsx
   with native Excel =HYPERLINK(...) formulas.
4. Monthly & Yearly Returns Heatmaps, Capital Growth Graphs, and Interactive Date Hover Equity HTML.
"""

import os
import sys
import math
import shutil
from pathlib import Path
import pandas as pd
import numpy as np

# Force UTF-8 encoding for stdout on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

DEFAULT_REPORTS_DIR = BASE_DIR / "Reports"
DEFAULT_PLOTS_DIR = BASE_DIR / "Plots" / "swing_statement_trades"
DATA_DAILY_DIR = BASE_DIR / "data_daily"

from swing_strategy.strategy_engine import prepare_swing_strategy_dataset, run_swing_strategy_ml_model
from swing_strategy.plotter import plot_swing_trade_chart
from swing_strategy.visualizer import generate_all_visualizations
from src.analysis.indian_brokerage_calculator import calculate_indian_trade_charges

# Pre-scan data_daily directory once for 100x disk lookup speedup
_DAILY_CSV_MAP = {}
_STOCK_OPEN_MAP = {}
_STOCK_C3_OPEN_MAP = {}

def get_daily_csv_map():
    global _DAILY_CSV_MAP
    if not _DAILY_CSV_MAP:
        _DAILY_CSV_MAP = {f.name: f for f in DATA_DAILY_DIR.glob("*.csv")}
    return _DAILY_CSV_MAP

def load_stock_maps(ticker: str):
    if ticker in _STOCK_OPEN_MAP:
        return _STOCK_OPEN_MAP[ticker], _STOCK_C3_OPEN_MAP[ticker]

    csv_map = get_daily_csv_map()
    clean_sym = ticker.replace(".NS", "")
    
    cand_names = [f"{ticker}_1d.csv", f"{ticker}.csv", f"{clean_sym}_1d.csv", f"{clean_sym}.csv"]
    
    open_map = {}
    c3_open_map = {}

    for name in cand_names:
        if name in csv_map:
            cand_path = csv_map[name]
            try:
                df = pd.read_csv(cand_path, usecols=["Date", "Open"])
                df["Date_Str"] = pd.to_datetime(df["Date"]).dt.strftime("%Y-%m-%d")
                df = df.sort_values("Date_Str").reset_index(drop=True)
                df["C3_Open"] = df["Open"].shift(-1)
                
                open_map = dict(zip(df["Date_Str"], df["Open"].astype(float)))
                c3_open_map = dict(zip(df["Date_Str"], df["C3_Open"].dropna().astype(float)))
            except Exception:
                pass
            break

    _STOCK_OPEN_MAP[ticker] = open_map
    _STOCK_C3_OPEN_MAP[ticker] = c3_open_map
    return open_map, c3_open_map


def get_c3_open_instant(ticker: str, c2_date_str: str) -> float:
    _, c3_map = load_stock_maps(ticker)
    return c3_map.get(c2_date_str)


def get_exit_open_instant(ticker: str, exit_date_str: str) -> float:
    open_map, _ = load_stock_maps(ticker)
    return open_map.get(exit_date_str)


def generate_swing_strategy_statement(
    initial_deposit: float = 100000.0,
    max_charts_to_generate: int = 50,
    max_risk_per_trade: float = 1000.0,
    df_acc: pd.DataFrame = None,
    custom_reports_dir: Path = None,
    custom_plots_dir: Path = None,
    exp_name: str = None
):
    if custom_reports_dir is None:
        custom_reports_dir = DEFAULT_REPORTS_DIR
    if custom_plots_dir is None:
        custom_plots_dir = DEFAULT_PLOTS_DIR

    trade_charts_dir = custom_plots_dir / "trade_charts" if exp_name else custom_plots_dir

    os.makedirs(custom_reports_dir, exist_ok=True)
    os.makedirs(trade_charts_dir, exist_ok=True)

    print(f"=== Swing Strategy Statement Generator [{exp_name or 'Default'}] (Initial Deposit: Rs {initial_deposit:,.2f}, Max Risk Cap: Rs {max_risk_per_trade:,.2f}) ===", flush=True)

    # 1. Load dataset & run walk-forward predictions if not provided
    if df_acc is None:
        df_6c = prepare_swing_strategy_dataset()
        df_acc = run_swing_strategy_ml_model(df_6c, probability_threshold=0.42)

    if "ML_Prediction" in df_acc.columns:
        df_acc = df_acc[df_acc["ML_Prediction"] == "Enter"].copy()

    df_acc["C2_Date"] = pd.to_datetime(df_acc["C2_Date"])
    df_acc = df_acc.sort_values("C2_Date").reset_index(drop=True)

    # 2. Chronological Simulation Setup
    current_balance = initial_deposit
    peak_balance = initial_deposit
    max_drawdown_pct = 0.0
    
    statement_rows = []
    daily_equity_rows = []
    
    # Add Initial Deposit Statement Entry
    deposit_date = "2010-01-01"
    statement_rows.append({
        "Transaction_ID": 1,
        "Trade_ID": 0,
        "Type": "DEPOSIT",
        "Date": deposit_date,
        "Ticker": "N/A",
        "Liquidity_Source": "N/A",
        "Support_Price": 0.0,
        "Quantity": 0,
        "Price": 0.0,
        "Total_Spend": 0.0,
        "Gross_PnL": 0.0,
        "Statutory_Taxes": 0.0,
        "Net_PnL": 0.0,
        "Return_Pct": 0.0,
        "Balance": current_balance,
        "Active_Position_Count": 0,
        "Target_RR_Mode": "N/A",
        "Outcome": "DEPOSIT",
        "Chart_PNG_URI": "N/A"
    })

    open_positions = []
    executed_trades_count = 0
    winning_trades_count = 0
    losing_trades_count = 0
    total_taxes_paid = 0.0
    gap_up_entries_count = 0

    trades_by_date = {}
    for r_dict in df_acc.to_dict("records"):
        trades_by_date.setdefault(r_dict["C2_Date"], []).append(r_dict)

    min_dt = df_acc["C2_Date"].min()
    max_exit = max(pd.to_datetime(df_acc["Exit_Date_1to2"]).max(), pd.to_datetime(df_acc["Exit_Date_1to3"]).max())
    all_days = pd.date_range(min_dt, max_exit, freq="D")

    tx_counter = 2

    tf_rank = {"Yearly": 3, "Monthly": 2, "Weekly": 1}
    nifty_rank = {"Nifty 50": 4, "Nifty 100": 3, "Nifty 250": 2, "Other": 1}

    for curr_dt in all_days:
        day_start_balance = current_balance
        day_net_pnl = 0.0

        # STEP 1: ENTRIES FIRST using cash available at the START of curr_dt
        if curr_dt in trades_by_date:
            candidates = trades_by_date[curr_dt]
            candidates.sort(key=lambda x: (-tf_rank.get(x.get("Liquidity_Type_1to2", "Weekly"), 0), -nifty_rank.get(x.get("Index_Membership_1to2", "Other"), 0)))

            allocated_cap = sum(p["Total_Spend"] for p in open_positions)
            available_cap = max(0.0, current_balance - allocated_cap)

            if available_cap > 0:
                for cand in candidates:
                    base_entry_p = cand["Entry_Price_1to2"]
                    rr_choice = cand.get("ML_RR_Choice", "1:2")

                    if rr_choice == "1:3":
                        sl_p = cand["SL_Price_1to3"]
                        tp_p = cand["Target_Price_1to3"]
                        exit_dt_val = cand["Exit_Date_1to3"]
                        outcome_val = cand["Outcome_1to3"]
                    else:
                        sl_p = cand["SL_Price_1to2"]
                        tp_p = cand["Target_Price_1to2"]
                        exit_dt_val = cand["Exit_Date_1to2"]
                        outcome_val = cand["Outcome_1to2"]

                    # Position Sizing based on max_risk_per_trade
                    risk_per_share = max(0.05, base_entry_p - sl_p)
                    qty = max(1, int(max_risk_per_trade / risk_per_share))
                    pos_val = base_entry_p * qty

                    if pos_val > available_cap or qty <= 0:
                        continue

                    c2_dt_str = curr_dt.strftime("%Y-%m-%d")
                    c3_open = get_c3_open_instant(cand["Ticker"], c2_dt_str)
                    if c3_open and c3_open > base_entry_p:
                        actual_entry_p = round(c3_open * 1.002, 2)  # Buy at 0.2% above C3 Open
                        gap_up_entries_count += 1
                    else:
                        actual_entry_p = base_entry_p

                    pos_val = actual_entry_p * qty

                    allocated_cap += pos_val
                    available_cap -= pos_val
                    bal_after_buy = current_balance - allocated_cap

                    trade_id = executed_trades_count + len(open_positions) + 1

                    pos_info = {
                        "Trade_ID": trade_id,
                        "Ticker": cand["Ticker"],
                        "Liquidity_Type": cand.get("Liquidity_Type_1to2", "Support Level"),
                        "Support_Price": cand["Support_Price"],
                        "C1_Date": cand["C1_Date"],
                        "C2_Date": cand["C2_Date"],
                        "Entry_Price": actual_entry_p,
                        "SL_Price": sl_p,
                        "Target_Price": tp_p,
                        "Exit_Date": exit_dt_val,
                        "Exit_Price": tp_p if outcome_val == "Success" else sl_p,
                        "Quantity": qty,
                        "Total_Spend": pos_val,
                        "Balance_After_Buy": bal_after_buy,
                        "ML_RR_Choice": rr_choice,
                        "Outcome": outcome_val
                    }
                    open_positions.append(pos_info)

                    statement_rows.append({
                        "Transaction_ID": tx_counter,
                        "Trade_ID": trade_id,
                        "Type": "BUY (ENTRY)",
                        "Date": curr_dt.strftime("%Y-%m-%d"),
                        "Ticker": cand["Ticker"],
                        "Liquidity_Source": cand.get("Liquidity_Type_1to2", "Support Level"),
                        "Support_Price": cand["Support_Price"],
                        "Quantity": qty,
                        "Price": actual_entry_p,
                        "Total_Spend": pos_val,
                        "Gross_PnL": 0.0,
                        "Statutory_Taxes": 0.0,
                        "Net_PnL": 0.0,
                        "Return_Pct": 0.0,
                        "Balance": bal_after_buy,
                        "Active_Position_Count": len(open_positions),
                        "Target_RR_Mode": rr_choice,
                        "Outcome": "OPEN",
                        "Chart_PNG_URI": "N/A"
                    })
                    tx_counter += 1

        # STEP 2: EXITS SECOND for open trades on curr_dt
        closed_positions = []
        for pos in open_positions:
            ex_dt = pd.to_datetime(pos["Exit_Date"])
            if curr_dt >= ex_dt:
                closed_positions.append(pos)

        for pos in closed_positions:
            open_positions.remove(pos)
            qty = pos["Quantity"]
            entry_p = pos["Entry_Price"]
            tot_spend = pos["Total_Spend"]
            target_p = pos["Target_Price"]
            sl_p = pos["SL_Price"]
            outcome_val = pos["Outcome"]

            exit_date_str = ex_dt.strftime("%Y-%m-%d")
            exit_open = get_exit_open_instant(pos["Ticker"], exit_date_str)

            if outcome_val == "Success":
                if exit_open and exit_open > target_p:
                    exit_p = round(exit_open * 0.999, 2)
                else:
                    exit_p = target_p
            else:
                if exit_open and exit_open < sl_p:
                    exit_p = round(exit_open * 0.999, 2)
                else:
                    exit_p = sl_p

            ch = calculate_indian_trade_charges(entry_p, exit_p, qty, flat_brokerage_per_order=0.0)
            gross_pnl = ch["gross_pnl"]
            taxes = ch["total_charges"]
            net_pnl = ch["net_pnl"]
            ret_pct = (net_pnl / tot_spend * 100.0) if tot_spend > 0 else 0.0

            current_balance += net_pnl
            day_net_pnl += net_pnl

            if current_balance > peak_balance:
                peak_balance = current_balance
            dd = (peak_balance - current_balance) / peak_balance * 100.0 if peak_balance > 0 else 0.0
            if dd > max_drawdown_pct:
                max_drawdown_pct = dd

            executed_trades_count += 1
            if net_pnl >= 0:
                winning_trades_count += 1
            else:
                losing_trades_count += 1
            total_taxes_paid += taxes

            chart_uri = "N/A"
            if executed_trades_count <= max_charts_to_generate:
                t_record = {**pos, "Exit_Price": exit_p, "Net_PnL": net_pnl, "Balance_After_Exit": current_balance}
                chart_uri = plot_swing_trade_chart(t_record, trade_charts_dir)

            statement_rows.append({
                "Transaction_ID": tx_counter,
                "Trade_ID": pos["Trade_ID"],
                "Type": "SELL (EXIT)",
                "Date": exit_date_str,
                "Ticker": pos["Ticker"],
                "Liquidity_Source": pos["Liquidity_Type"],
                "Support_Price": pos["Support_Price"],
                "Quantity": qty,
                "Price": exit_p,
                "Total_Spend": tot_spend,
                "Gross_PnL": gross_pnl,
                "Statutory_Taxes": taxes,
                "Net_PnL": net_pnl,
                "Return_Pct": ret_pct,
                "Balance": current_balance,
                "Active_Position_Count": len(open_positions),
                "Target_RR_Mode": pos["ML_RR_Choice"],
                "Outcome": pos["Outcome"],
                "Chart_PNG_URI": chart_uri
            })
            tx_counter += 1

        daily_ret = (day_net_pnl / day_start_balance * 100.0) if day_start_balance > 0 else 0.0
        daily_equity_rows.append({
            "Date": curr_dt.strftime("%Y-%m-%d"),
            "Balance": current_balance,
            "Active_Positions": len(open_positions),
            "Daily_PnL": day_net_pnl,
            "Daily_Return_Pct": daily_ret
        })

    df_stmt = pd.DataFrame(statement_rows)
    df_daily_equity = pd.DataFrame(daily_equity_rows)

    net_return_pct = ((current_balance - initial_deposit) / initial_deposit * 100.0) if initial_deposit > 0 else 0.0
    cagr_pct = (((current_balance / initial_deposit) ** (1 / 16.0) - 1) * 100.0) if (initial_deposit > 0 and current_balance > 0) else 0.0
    win_rate_pct = (winning_trades_count / executed_trades_count * 100.0) if executed_trades_count > 0 else 0.0

    print(f"[{exp_name or 'Default'}] Final Balance: Rs {current_balance:,.2f} | Net Return: +{net_return_pct:,.2f}% | CAGR: {cagr_pct:.2f}% | Win Rate: {win_rate_pct:.2f}% ({executed_trades_count:,} Trades)", flush=True)

    excel_path = custom_reports_dir / "Swing_Strategy_Account_Statement.xlsx"
    csv_path = custom_reports_dir / "Swing_Strategy_Account_Statement.csv"

    summary_dict = {
        "Exp_Name": exp_name or "Default",
        "Initial_Capital": initial_deposit,
        "Max_Risk_Cap": max_risk_per_trade,
        "Final_Equity": current_balance,
        "Total_Net_Return_Pct": net_return_pct,
        "CAGR_Pct": cagr_pct,
        "Executed_Trades": executed_trades_count,
        "Gap_Up_Fills": gap_up_entries_count,
        "Winning_Trades": winning_trades_count,
        "Losing_Trades": losing_trades_count,
        "Win_Rate_Pct": win_rate_pct,
        "Max_Drawdown_Pct": max_drawdown_pct,
        "Total_Taxes_Paid": total_taxes_paid
    }

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws_stmt = wb.active
        ws_stmt.title = "Account Statement"

        headers = [
            "Tx ID", "Trade ID", "Type", "Date", "Ticker", "Liquidity Source", "Support Price (Rs)",
            "Quantity", "Price (Rs)", "Total Spend (Rs)", "Gross PnL (Rs)", "Taxes/Charges (Rs)",
            "Net PnL (Rs)", "Return (%)", "Account Balance (Rs)", "Active Position Count", "Target RR Mode", "Outcome", "Trade Graph Link"
        ]
        ws_stmt.append(headers)

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_num in range(1, len(headers) + 1):
            cell = ws_stmt.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        buy_fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
        sell_win_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
        sell_loss_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

        for row_idx, r in df_stmt.iterrows():
            row_num = row_idx + 2
            tx_type = r["Type"]
            pnl = r["Net_PnL"]
            
            row_data = [
                r["Transaction_ID"], r["Trade_ID"], r["Type"], r["Date"], r["Ticker"],
                r["Liquidity_Source"], r["Support_Price"], r["Quantity"], r["Price"],
                r["Total_Spend"], r["Gross_PnL"], r["Statutory_Taxes"], r["Net_PnL"],
                r["Return_Pct"], r["Balance"], r["Active_Position_Count"], r["Target_RR_Mode"], r["Outcome"], r["Chart_PNG_URI"]
            ]
            
            ws_stmt.append(row_data)

            fill = buy_fill if tx_type.startswith("BUY") else (sell_win_fill if pnl >= 0 else sell_loss_fill)
            for col_idx in range(1, len(headers) + 1):
                cell = ws_stmt.cell(row=row_num, column=col_idx)
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")

                if col_idx == 19 and r["Chart_PNG_URI"] != "N/A":
                    uri = r["Chart_PNG_URI"]
                    cell.value = f'=HYPERLINK("{uri}", "View Plot Chart (PNG)")'
                    cell.hyperlink = uri
                    cell.font = Font(name="Calibri", size=10, color="2563EB", underline="single")

        ws_sum = wb.create_sheet(title="Performance Summary")
        ws_sum.append(["Performance Metric", "Strategy Result"])
        ws_sum.cell(row=1, column=1).font = header_font
        ws_sum.cell(row=1, column=1).fill = header_fill
        ws_sum.cell(row=1, column=2).font = header_font
        ws_sum.cell(row=1, column=2).fill = header_fill

        for k_name, val in summary_dict.items():
            ws_sum.append([k_name, str(val)])

        wb.save(excel_path)

    except Exception as e:
        print(f"Excel styling export warning: {e}", flush=True)

    df_stmt.to_csv(csv_path, index=False)

    # Generate Visualizations Suite for this Experiment
    generate_all_visualizations(
        df_daily_equity,
        output_plots_dir=custom_plots_dir,
        output_reports_dir=custom_reports_dir,
        exp_title=f"Swing Strategy ({exp_name or 'Default'})"
    )

    return summary_dict


if __name__ == "__main__":
    generate_swing_strategy_statement(initial_deposit=100000.0, max_charts_to_generate=50, max_risk_per_trade=1000.0)
